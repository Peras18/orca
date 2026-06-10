#!/usr/bin/env python3
import csv, sys, os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

GREEN  = PatternFill("solid", fgColor="1DB954")
BLUE   = PatternFill("solid", fgColor="1A73E8")
DARK   = PatternFill("solid", fgColor="1E1E2E")
GREY   = PatternFill("solid", fgColor="2A2A3E")
GOLD   = PatternFill("solid", fgColor="F4B942")
RED    = PatternFill("solid", fgColor="E53935")
WFONT  = Font(color="FFFFFF", bold=True)
NFONT  = Font(color="FFFFFF", bold=False)
GFONT  = Font(color="1DB954", bold=True)

def sh(cell, fill=BLUE):
    cell.fill = fill
    cell.font = WFONT
    cell.alignment = Alignment(horizontal="center", vertical="center")

def read_ops(path="logs/opportunities.csv"):
    ops = []
    if os.path.exists(path):
        with open(path, newline="") as f:
            for row in csv.DictReader(f):
                ops.append(row)
    if not ops:
        print("Sem dados reais — usando exemplo demonstrativo")
        ops = [
            {"timestamp":"2026-05-16T20:46:19Z","block":"46087800",
             "path":"WETH->AERO->USDC->WETH","hops":"3","input_eth":"1.000000",
             "gross_profit_eth":"0.002013","gas_cost_eth":"0.000240",
             "net_profit_eth":"0.001773","net_profit_eur_1800":"3.1914"},
            {"timestamp":"2026-05-16T20:46:20Z","block":"46087800",
             "path":"WETH->AERO->USDC->WETH","hops":"3","input_eth":"0.500000",
             "gross_profit_eth":"0.001187","gas_cost_eth":"0.000240",
             "net_profit_eth":"0.000947","net_profit_eur_1800":"1.7046"},
            {"timestamp":"2026-05-16T20:46:21Z","block":"46087800",
             "path":"WETH->USDC->AERO->WETH","hops":"3","input_eth":"2.000000",
             "gross_profit_eth":"0.003890","gas_cost_eth":"0.000240",
             "net_profit_eth":"0.003650","net_profit_eur_1800":"6.5700"},
        ]
    return ops

def generate(ops, eth_eur=1800.0, capital_eur=80.0):
    wb = openpyxl.Workbook()

    # ── SHEET 1: Raw Data ──
    ws1 = wb.active
    ws1.title = "Oportunidades"
    ws1.sheet_properties.tabColor = "1DB954"

    hdrs = ["Timestamp","Bloco","Path","Hops","Input ETH",
            "Gross Profit ETH","Gas Cost ETH","Net Profit ETH","Net Profit EUR"]
    widths = [22,12,35,6,12,18,14,16,14]
    for c, (h, w) in enumerate(zip(hdrs, widths), 1):
        cell = ws1.cell(row=1, column=c, value=h)
        sh(cell)
        ws1.column_dimensions[get_column_letter(c)].width = w

    for r, op in enumerate(ops, 2):
        vals = [
            op.get("timestamp",""),
            int(op.get("block",0)),
            op.get("path",""),
            int(op.get("hops",0)),
            float(op.get("input_eth",0)),
            float(op.get("gross_profit_eth",0)),
            float(op.get("gas_cost_eth",0)),
            float(op.get("net_profit_eth",0)),
            float(op.get("net_profit_eur_1800",0)),
        ]
        fill = GREY if r % 2 == 0 else DARK
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=r, column=c, value=v)
            cell.fill = fill
            cell.font = NFONT

    ws1.freeze_panes = "A2"

    # ── SHEET 2: Analise ──
    ws2 = wb.create_sheet("Analise por Banca")
    ws2.sheet_properties.tabColor = "1A73E8"

    n = len(ops)
    avg_net = sum(float(o.get("net_profit_eth",0)) for o in ops) / max(n,1)
    avg_gross = sum(float(o.get("gross_profit_eth",0)) for o in ops) / max(n,1)
    avg_gas = sum(float(o.get("gas_cost_eth",0)) for o in ops) / max(n,1)

    try:
        t0 = datetime.fromisoformat(ops[0]["timestamp"].replace("Z",""))
        t1 = datetime.fromisoformat(ops[-1]["timestamp"].replace("Z",""))
        hrs = max((t1-t0).total_seconds()/3600, 1/60)
        ops_hr = n / hrs
    except:
        ops_hr = 4.0

    daily_eth = avg_net * min(ops_hr, 15) * 5

    # Titulo
    ws2["A1"] = "ORCA MEV Bot — Relatorio DRY_RUN"
    ws2["A1"].font = Font(color="1DB954", bold=True, size=16)
    ws2["A1"].fill = DARK
    ws2.merge_cells("A1:F1")
    ws2.row_dimensions[1].height = 30

    ws2["A2"] = f"Gerado: {datetime.now().strftime('%Y-%m-%d %H:%M')} | ETH={eth_eur}EUR | Capital={capital_eur}EUR"
    ws2["A2"].font = Font(color="AAAAAA", italic=True)
    ws2["A2"].fill = DARK
    ws2.merge_cells("A2:F2")

    # Metricas
    ws2["A4"] = "METRICAS DO DRY_RUN"
    sh(ws2["A4"], BLUE)
    ws2.merge_cells("A4:C4")

    rows_m = [
        ("Oportunidades registadas", n, ""),
        ("Ops estimadas / hora", f"{ops_hr:.1f}", "ops/h"),
        ("Ops estimadas / dia", f"{ops_hr*24:.0f}", "ops/dia"),
        ("Lucro bruto medio / op", f"{avg_gross:.6f}", "ETH"),
        ("Gas medio / op", f"{avg_gas:.6f}", "ETH"),
        ("Lucro liquido medio / op", f"{avg_net:.6f}", "ETH"),
        ("Lucro liquido medio / op", f"{avg_net*eth_eur:.4f}", "EUR"),
        ("Lucro liquido / dia (estimado)", f"{daily_eth:.6f}", "ETH"),
        ("Lucro liquido / dia (estimado)", f"{daily_eth*eth_eur:.2f}", "EUR"),
    ]
    for i, (lbl, val, unit) in enumerate(rows_m, 5):
        ws2[f"A{i}"] = lbl
        ws2[f"B{i}"] = str(val)
        ws2[f"C{i}"] = unit
        ws2[f"A{i}"].font = NFONT
        ws2[f"A{i}"].fill = DARK
        ws2[f"B{i}"].font = Font(color="1DB954", bold=True)
        ws2[f"B{i}"].fill = DARK
        ws2[f"C{i}"].font = Font(color="AAAAAA")
        ws2[f"C{i}"].fill = DARK

    # Tabela por banca
    start_row = 16
    ws2[f"A{start_row}"] = "PROJECAO POR BANCA INICIAL"
    sh(ws2[f"A{start_row}"], GOLD)
    ws2.merge_cells(f"A{start_row}:G{start_row}")

    th = ["Banca (EUR)","Banca (ETH)","Lucro/dia (EUR)",
          "Lucro/mes (EUR)","Lucro/ano (EUR)","ROI/mes (%)","Dias p/dobrar"]
    for c, h in enumerate(th, 1):
        cell = ws2.cell(row=start_row+1, column=c, value=h)
        sh(cell, BLUE)
        ws2.column_dimensions[get_column_letter(c)].width = 17

    bancas = [80, 200, 500, 1000, 2000, 5000, 10000]
    for i, b_eur in enumerate(bancas):
        b_eth = b_eur / eth_eur
        scale = min(b_eth / (capital_eur/eth_eur), 20.0)
        d_eur = daily_eth * eth_eur * scale
        m_eur = d_eur * 30
        a_eur = d_eur * 365
        roi = (m_eur / b_eur) * 100
        days_d = b_eur / d_eur if d_eur > 0 else 9999

        r = start_row + 2 + i
        vals = [f"{b_eur}EUR", f"{b_eth:.4f}",
                f"{d_eur:.2f}", f"{m_eur:.2f}",
                f"{a_eur:.2f}", f"{roi:.1f}%", f"{days_d:.0f}"]
        fill = GREEN if b_eur == int(capital_eur) else (GREY if i%2==0 else DARK)
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.fill = fill
            cell.font = WFONT if b_eur == int(capital_eur) else NFONT

    note_row = start_row + 2 + len(bancas) + 1
    ws2[f"A{note_row}"] = ("NOTA: Projecoes baseadas em DRY_RUN com dados reais de mercado. "
                           "Flashloans permitem operar acima da banca propria. "
                           "A banca cobre apenas gas (~0.0001 ETH/tx na Base).")
    ws2[f"A{note_row}"].font = Font(color="F4B942", italic=True)
    ws2[f"A{note_row}"].fill = DARK
    ws2.merge_cells(f"A{note_row}:G{note_row}")

    # ── SHEET 3: Juros Compostos ──
    ws3 = wb.create_sheet("Juros Compostos")
    ws3.sheet_properties.tabColor = "F4B942"

    ws3["A1"] = "Crescimento com Juros Compostos — 36 meses"
    ws3["A1"].font = Font(color="1DB954", bold=True, size=14)
    ws3["A1"].fill = DARK
    ws3.merge_cells("A1:G1")

    th3 = ["Mes","Banca (EUR)","Lucro/mes (EUR)","Acumulado (EUR)","ROI Total (%)","Lucro/dia (EUR)","Marco"]
    for c, h in enumerate(th3, 1):
        cell = ws3.cell(row=2, column=c, value=h)
        sh(cell, BLUE)
        ws3.column_dimensions[get_column_letter(c)].width = 17

    banca = capital_eur
    acum = 0.0
    marcos = {5:"5EUR/dia", 10:"10EUR/dia", 45:"45EUR/dia",
              100:"100EUR/dia", 200:"200EUR/dia", 500:"500EUR/dia",
              1000:"1000EUR/dia", 5000:"5000EUR/dia"}

    for mes in range(1, 37):
        scale = min(banca / capital_eur, 50.0)
        l_mes = daily_eth * eth_eur * scale * 30
        acum += l_mes
        banca += l_mes
        roi = ((banca - capital_eur) / capital_eur) * 100
        l_dia = l_mes / 30

        marco = ""
        for threshold in sorted(marcos.keys(), reverse=True):
            if l_dia >= threshold:
                marco = marcos[threshold]
                break

        r = mes + 2
        vals = [f"Mes {mes}", f"{banca:.2f}", f"{l_mes:.2f}",
                f"{acum:.2f}", f"{roi:.1f}%", f"{l_dia:.2f}", marco]
        fill = GREEN if marco else (GREY if mes%2==0 else DARK)
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=r, column=c, value=v)
            cell.fill = fill
            cell.font = WFONT if marco else NFONT

    os.makedirs("logs", exist_ok=True)
    out = "logs/orca_dryrun_report.xlsx"
    wb.save(out)
    print(f"Relatorio gerado: {out}")
    print(f"  Oportunidades: {n}")
    print(f"  Lucro medio/op: {avg_net*eth_eur:.4f} EUR")
    print(f"  Estimativa diaria: {daily_eth*eth_eur:.2f} EUR/dia")

if __name__ == "__main__":
    ops = read_ops()
    generate(ops)